from __future__ import annotations

import argparse
import io
import json
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, MutableMapping, Optional, Sequence, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REST_LABEL = "休"
NIGHT_SHIFT_CODES = {"NA", "NB", "NC"}
NIGHT_SHIFT_NAMES = {"夜勤A", "夜勤B", "夜勤C"}
DEMAND_SLOTS = ["7-9", "9-15", "16-18", "18-21", "21-23", "0-7"]


def _load_shift_name_map() -> Dict[str, str]:
    catalog_path = Path(__file__).with_name("shifts_catalog.json")
    mapping: Dict[str, str] = {}
    try:
        raw = json.loads(catalog_path.read_text(encoding="utf-8"))
    except FileNotFoundError:
        raw = []
    if isinstance(raw, list):
        for entry in raw:
            if not isinstance(entry, Mapping):
                continue
            code = entry.get("code")
            if not isinstance(code, str):
                continue
            name = entry.get("name")
            mapping[code.upper()] = str(name or code)
    mapping.setdefault("REST", REST_LABEL)
    mapping.setdefault("OFF", REST_LABEL)
    return mapping


SHIFT_NAME_MAP = _load_shift_name_map()


def _to_record(value: Any) -> Optional[MutableMapping[str, Any]]:
    if isinstance(value, MutableMapping):
        return value
    return None


def _normalize_shift(value: Any) -> str:
    if value is None:
        return REST_LABEL
    if isinstance(value, str):
        trimmed = value.strip()
        if not trimmed:
            return REST_LABEL
        if trimmed in {REST_LABEL, "休日", "休み"}:
            return REST_LABEL
        upper = trimmed.upper()
        if upper in SHIFT_NAME_MAP:
            return SHIFT_NAME_MAP[upper]
        if trimmed in SHIFT_NAME_MAP.values():
            return trimmed
        if upper in {"REST", "OFF", "HOLIDAY"}:
            return REST_LABEL
        return trimmed
    if isinstance(value, (int, float)):
        return str(value)
    return str(value)


def _infer_days(data: Mapping[str, Any]) -> int:
    candidates: List[int] = []

    direct_days = data.get("days")
    if isinstance(direct_days, int) and direct_days > 0:
        candidates.append(direct_days)

    meta = _to_record(data.get("meta"))
    if meta:
        value = meta.get("days")
        if isinstance(value, int) and value > 0:
            candidates.append(value)

    summary = _to_record(data.get("summary"))
    if summary:
        diagnostics = _to_record(summary.get("diagnostics"))
        if diagnostics:
            demand = _to_record(diagnostics.get("demand"))
            if demand:
                value = demand.get("days")
                if isinstance(value, int) and value > 0:
                    candidates.append(value)

    matrix_entries = data.get("matrix")
    if isinstance(matrix_entries, list):
        candidates.append(len(matrix_entries))
        for entry in matrix_entries:
            if isinstance(entry, Mapping):
                date = entry.get("date")
                if isinstance(date, int) and date > 0:
                    candidates.append(date)

    assignments = data.get("assignments")
    if isinstance(assignments, list):
        for entry in assignments:
            if isinstance(entry, Mapping):
                date = entry.get("date")
                if isinstance(date, int) and date > 0:
                    candidates.append(date)

    if not candidates:
        return 0
    return max(candidates)


def _infer_people_order(data: Mapping[str, Any]) -> List[str]:
    order: List[str] = []

    raw_order = data.get("peopleOrder")
    if isinstance(raw_order, Sequence) and not isinstance(raw_order, (str, bytes)):
        for value in raw_order:
            if isinstance(value, str) and value not in order:
                order.append(value)

    people = data.get("people")
    if not order and isinstance(people, Sequence) and not isinstance(people, (str, bytes)):
        for person in people:
            if isinstance(person, Mapping):
                pid = person.get("id")
                if isinstance(pid, str) and pid not in order:
                    order.append(pid)

    input_echo = _to_record(data.get("inputEcho"))
    if not order and input_echo:
        echo_people = input_echo.get("people")
        if isinstance(echo_people, Sequence) and not isinstance(echo_people, (str, bytes)):
            for person in echo_people:
                if isinstance(person, Mapping):
                    pid = person.get("id")
                    if isinstance(pid, str) and pid not in order:
                        order.append(pid)

    matrix_entries = data.get("matrix")
    if isinstance(matrix_entries, list):
        for entry in matrix_entries:
            if not isinstance(entry, Mapping):
                continue
            shifts = entry.get("shifts")
            if isinstance(shifts, Mapping):
                for pid in shifts.keys():
                    if isinstance(pid, str) and pid not in order:
                        order.append(pid)

    assignments = data.get("assignments")
    if isinstance(assignments, list):
        for entry in assignments:
            if not isinstance(entry, Mapping):
                continue
            staff_id = entry.get("staffId") or entry.get("id")
            if isinstance(staff_id, str) and staff_id not in order:
                order.append(staff_id)

    return order


def _build_schedule_matrix(
    data: Mapping[str, Any],
    days: int,
    staff: Sequence[str],
) -> List[Dict[str, str]]:
    staff_list = list(staff)
    matrix: List[Dict[str, str]] = []
    default_row = {pid: REST_LABEL for pid in staff_list}

    matrix_entries: Dict[int, Dict[str, str]] = {}
    raw_matrix = data.get("matrix")
    if isinstance(raw_matrix, list):
        for entry in raw_matrix:
            if not isinstance(entry, Mapping):
                continue
            day = entry.get("date")
            if not isinstance(day, int) or day < 1:
                continue
            shifts = entry.get("shifts")
            if isinstance(shifts, Mapping):
                matrix_entries[day] = {
                    pid: _normalize_shift(shifts.get(pid)) for pid in staff_list
                }

    for day in range(1, days + 1):
        row = dict(default_row)
        if day in matrix_entries:
            row.update(matrix_entries[day])
        matrix.append(row)

    assignments = data.get("assignments")
    if isinstance(assignments, list):
        for entry in assignments:
            if not isinstance(entry, Mapping):
                continue
            if isinstance(entry.get("shifts"), Sequence) and isinstance(entry.get("id"), str):
                staff_id = entry["id"]  # type: ignore[index]
                if staff_id not in staff_list:
                    continue
                shifts_seq = entry.get("shifts")
                if isinstance(shifts_seq, Sequence):
                    for index, shift_value in enumerate(shifts_seq, start=1):
                        if 1 <= index <= days:
                            matrix[index - 1][staff_id] = _normalize_shift(shift_value)
                continue

            day = entry.get("date")
            staff_id = entry.get("staffId")
            shift_value = entry.get("shift")
            if (
                isinstance(day, int)
                and isinstance(staff_id, str)
                and 1 <= day <= days
                and staff_id in matrix[day - 1]
            ):
                matrix[day - 1][staff_id] = _normalize_shift(shift_value)

    return matrix


def _auto_adjust_column_width(worksheet) -> None:
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            text = str(value)
            if len(text) > max_length:
                max_length = len(text)
        adjusted_width = max(10, min(40, max_length + 2))
        worksheet.column_dimensions[column_letter].width = adjusted_width


def _apply_shift_formatting(worksheet, staff_count: int, days: int) -> None:
    header_font = Font(bold=True)
    for cell in worksheet[2]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    data_alignment = Alignment(horizontal="center", vertical="center")
    rest_fill = PatternFill(fill_type="solid", start_color="E5E7EB", end_color="E5E7EB")
    night_fill = PatternFill(fill_type="solid", start_color="DBEAFE", end_color="DBEAFE")

    for row in worksheet.iter_rows(min_row=3, max_row=2 + days, min_col=2, max_col=1 + staff_count):
        for cell in row:
            cell.alignment = data_alignment
            value = cell.value
            if not isinstance(value, str):
                continue
            text = value.strip()
            if not text:
                continue
            if text == REST_LABEL:
                cell.fill = rest_fill
            else:
                upper = text.upper()
                if upper in NIGHT_SHIFT_CODES or text in NIGHT_SHIFT_NAMES or "夜勤" in text:
                    cell.fill = night_fill

    worksheet.freeze_panes = "B3"


def _build_shortage_frame(data: Mapping[str, Any], days: int) -> Optional[pd.DataFrame]:
    summary = _to_record(data.get("summary"))
    if not summary:
        return None
    shortage_entries = summary.get("shortage")
    if not isinstance(shortage_entries, list):
        return None

    table = {slot: {day: 0 for day in range(1, days + 1)} for slot in DEMAND_SLOTS}
    has_value = False
    for entry in shortage_entries:
        if not isinstance(entry, Mapping):
            continue
        day = entry.get("date")
        slot = entry.get("slot")
        lack = entry.get("lack")
        if isinstance(day, int) and isinstance(slot, str) and slot in table:
            if isinstance(lack, (int, float)):
                table[slot][day] = int(lack)
                if int(lack) != 0:
                    has_value = True
    if not has_value:
        return None

    rows = []
    for slot in DEMAND_SLOTS:
        row = {"時間帯": slot}
        row.update(table[slot])
        rows.append(row)
    return pd.DataFrame(rows)


def _build_need_template_frame(data: Mapping[str, Any]) -> Optional[pd.DataFrame]:
    candidates: List[Mapping[str, Any]] = []
    need_template = data.get("needTemplate")
    if isinstance(need_template, Mapping):
        candidates.append(need_template)

    input_echo = _to_record(data.get("inputEcho"))
    if input_echo:
        echo_template = input_echo.get("needTemplate")
        if isinstance(echo_template, Mapping):
            candidates.append(echo_template)

    summary = _to_record(data.get("summary"))
    if summary:
        diagnostics = _to_record(summary.get("diagnostics"))
        if diagnostics:
            demand = _to_record(diagnostics.get("demand"))
            if demand:
                template = demand.get("needTemplate")
                if isinstance(template, Mapping):
                    candidates.append(template)

    template_map: Optional[Mapping[str, Any]] = None
    for candidate in candidates:
        if candidate:
            template_map = candidate
            break
    if not template_map:
        return None

    rows = []
    for day_type in sorted(template_map.keys()):
        entry = template_map.get(day_type)
        if not isinstance(entry, Mapping):
            continue
        row = {"dayType": day_type}
        for slot in DEMAND_SLOTS:
            value = entry.get(slot)
            if isinstance(value, (int, float)):
                row[slot] = int(value)
            else:
                row[slot] = value
        rows.append(row)
    if not rows:
        return None
    return pd.DataFrame(rows)


def _translate_shift_codes(codes: Iterable[Any]) -> str:
    if isinstance(codes, (str, bytes)):
        return _normalize_shift(codes)
    names: List[str] = []
    for code in codes:
        names.append(_normalize_shift(code))
    return " / ".join([name for name in names if name])


def _format_sequence(values: Iterable[Any]) -> str:
    if isinstance(values, (str, bytes)):
        text = values.strip()
        return text
    formatted: List[str] = []
    for value in values:
        if value is None:
            continue
        text = str(value)
        if text:
            formatted.append(text)
    return " / ".join(formatted)


def _build_staff_conditions_frame(data: Mapping[str, Any]) -> Optional[pd.DataFrame]:
    people_sources: List[Sequence[Any]] = []
    raw_people = data.get("people")
    if isinstance(raw_people, Sequence) and not isinstance(raw_people, (str, bytes)):
        people_sources.append(raw_people)
    input_echo = _to_record(data.get("inputEcho"))
    if input_echo:
        echo_people = input_echo.get("people")
        if isinstance(echo_people, Sequence) and not isinstance(echo_people, (str, bytes)):
            people_sources.append(echo_people)

    people_list: Optional[Sequence[Any]] = None
    for candidate in people_sources:
        if candidate:
            people_list = candidate
            break
    if not people_list:
        return None

    rows = []
    for person in people_list:
        if not isinstance(person, Mapping):
            continue
        pid = person.get("id")
        if not isinstance(pid, str):
            continue
        can_work = person.get("canWork")
        fixed_off_weekdays = person.get("fixedOffWeekdays")
        fixed_off_dates = person.get("fixedOffDates")
        row = {
            "スタッフ": pid,
            "勤務可": (
                _translate_shift_codes(can_work)
                if isinstance(can_work, Iterable) and not isinstance(can_work, (str, bytes))
                else ""
            ),
            "固定休": (
                _format_sequence(fixed_off_weekdays)
                if isinstance(fixed_off_weekdays, Iterable) and not isinstance(fixed_off_weekdays, (str, bytes))
                else ""
            ),
            "固定休(日付)": (
                _format_sequence(fixed_off_dates)
                if isinstance(fixed_off_dates, Iterable) and not isinstance(fixed_off_dates, (str, bytes))
                else ""
            ),
            "週下限": person.get("weeklyMin"),
            "週上限": person.get("weeklyMax"),
            "月下限": person.get("monthlyMin"),
            "月上限": person.get("monthlyMax"),
            "連勤上限": person.get("consecMax"),
        }
        rows.append(row)
    if not rows:
        return None
    return pd.DataFrame(rows)


def _build_diagnostics_frame(data: Mapping[str, Any]) -> Optional[pd.DataFrame]:
    rows: List[Dict[str, Any]] = []

    infeasible = data.get("infeasible")
    if infeasible is not None:
        rows.append({"項目": "infeasible", "値": str(infeasible)})

    summary = _to_record(data.get("summary"))
    if summary:
        totals = _to_record(summary.get("totals"))
        if totals:
            for key, value in totals.items():
                rows.append({"項目": f"summary.totals.{key}", "値": value})
        diagnostics = _to_record(summary.get("diagnostics"))
        if diagnostics:
            demand = _to_record(diagnostics.get("demand"))
            if demand:
                total_need = demand.get("totalNeed")
                if total_need is not None:
                    rows.append({"項目": "summary.diagnostics.demand.totalNeed", "値": total_need})
                warnings = demand.get("warnings")
                if isinstance(warnings, list) and warnings:
                    joined = "; ".join(
                        str(item) for item in warnings if isinstance(item, (str, int, float))
                    )
                    if joined:
                        rows.append({"項目": "summary.diagnostics.demand.warnings", "値": joined})

    diagnostics = _to_record(data.get("diagnostics"))
    if diagnostics:
        var_counts = _to_record(diagnostics.get("var_counts"))
        if var_counts:
            for key, value in var_counts.items():
                rows.append({"項目": f"diagnostics.var_counts.{key}", "値": value})
        flags = _to_record(diagnostics.get("flags"))
        if flags:
            for key, value in flags.items():
                rows.append({"項目": f"diagnostics.flags.{key}", "値": value})
        warnings = diagnostics.get("warnings")
        if isinstance(warnings, list) and warnings:
            joined = "; ".join(
                str(item) for item in warnings if isinstance(item, (str, int, float))
            )
            if joined:
                rows.append({"項目": "diagnostics.warnings", "値": joined})
        log_output = diagnostics.get("logOutput")
        if isinstance(log_output, str) and log_output.strip():
            rows.append({"項目": "diagnostics.logOutput", "値": log_output.strip()})

    if not rows:
        return None
    return pd.DataFrame(rows)


def _extract_year_month(data: Mapping[str, Any]) -> Tuple[Optional[int], Optional[int]]:
    sources: List[Mapping[str, Any]] = []
    for candidate in (data, data.get("meta"), data.get("inputEcho")):
        record = _to_record(candidate)
        if record:
            sources.append(record)
            nested_meta = _to_record(record.get("meta"))
            if nested_meta:
                sources.append(nested_meta)
    for source in sources:
        year = source.get("year")
        month = source.get("month")
        if isinstance(year, int) and isinstance(month, int) and 1 <= month <= 12:
            return year, month
    return None, None


def build_default_filename(data: Mapping[str, Any], prefix: str = "shift-schedule") -> str:
    year, month = _extract_year_month(data)
    if isinstance(year, int) and isinstance(month, int):
        return f"{prefix}-{year}-{month:02d}.xlsx"
    return f"{prefix}.xlsx"


def _write_workbook(writer, data: Mapping[str, Any]) -> None:
    days = _infer_days(data)
    if days <= 0:
        raise ValueError("days could not be inferred from the provided JSON")
    staff = _infer_people_order(data)
    schedule_matrix = _build_schedule_matrix(data, days, staff)
    df = pd.DataFrame(schedule_matrix, index=range(1, days + 1), columns=staff)
    df.index.name = "日"

    df.to_excel(writer, sheet_name="シフト表", startrow=1)
    worksheet = writer.sheets["シフト表"]
    title = "シフト表"
    year, month = _extract_year_month(data)
    if isinstance(year, int) and isinstance(month, int):
        title = f"{year}/{month:02d} シフト表"
    worksheet["A1"] = title
    end_column = max(1, len(staff) + 1)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    worksheet["A1"].font = Font(size=14, bold=True)
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

    _apply_shift_formatting(worksheet, len(staff), days)
    _auto_adjust_column_width(worksheet)

    shortage_frame = _build_shortage_frame(data, days)
    if shortage_frame is not None:
        shortage_frame.to_excel(writer, sheet_name="不足サマリー", index=False)
        shortage_ws = writer.sheets["不足サマリー"]
        shortage_ws.freeze_panes = "B2"
        _auto_adjust_column_width(shortage_ws)

    need_template_frame = _build_need_template_frame(data)
    if need_template_frame is not None:
        need_template_frame.to_excel(writer, sheet_name="需要テンプレ", index=False)
        need_ws = writer.sheets["需要テンプレ"]
        need_ws.freeze_panes = "B2"
        _auto_adjust_column_width(need_ws)

    staff_frame = _build_staff_conditions_frame(data)
    if staff_frame is not None:
        staff_frame.to_excel(writer, sheet_name="スタッフ条件", index=False)
        staff_ws = writer.sheets["スタッフ条件"]
        staff_ws.freeze_panes = "B2"
        _auto_adjust_column_width(staff_ws)

    diagnostics_frame = _build_diagnostics_frame(data)
    if diagnostics_frame is not None:
        diagnostics_frame.to_excel(writer, sheet_name="診断", index=False)
        diag_ws = writer.sheets["診断"]
        diag_ws.freeze_panes = "B2"
        _auto_adjust_column_width(diag_ws)


def export_to_xlsx_bytes(data: Mapping[str, Any]) -> bytes:
    if not isinstance(data, Mapping):
        raise ValueError("Input data must be a mapping object")
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        _write_workbook(writer, data)
    return buffer.getvalue()


def export_to_xlsx_file(data: Mapping[str, Any], output_path: Path) -> None:
    if not isinstance(data, Mapping):
        raise ValueError("Input data must be a mapping object")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _write_workbook(writer, data)


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert solver output JSON into an Excel workbook.")
    parser.add_argument("--in", dest="infile", required=True, help="Path to solver output JSON file.")
    parser.add_argument("--out", dest="outfile", required=True, help="Path to write the Excel workbook.")
    args = parser.parse_args()

    in_path = Path(args.infile)
    out_path = Path(args.outfile)

    data = json.loads(in_path.read_text(encoding="utf-8"))
    if not isinstance(data, Mapping):
        raise SystemExit("Input JSON must be an object.")

    workbook_bytes = export_to_xlsx_bytes(data)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_bytes(workbook_bytes)
    print(f"wrote {out_path}")


if __name__ == "__main__":
    main()
